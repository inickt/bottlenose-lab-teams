import argparse
import csv
from dataclasses import dataclass
from getpass import getpass
from io import BytesIO
import os
import re
from typing import Optional, List, Tuple

from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook, load_workbook
from more_itertools import divide


@dataclass
class Team:
    team_id: int
    active: bool
    lab_section: Optional[int]
    member1: str
    member2: Optional[str]
    member3: Optional[str]

    def __lt__(self, other):
        return self.team_id < other.team_id

    @property
    def member_names(self) -> str:
        if self.member2 and self.member3:
            return f"{self.member3}, {self.member2}, and {self.member1}"
        if self.member2:
            return f"{self.member2} and {self.member1}"
        return self.member1

    @property
    def ms_team_channel(self) -> str:
        return f"Team {self.team_id} - {self.member_names}"


def without_email(name: str) -> str:
    return re.sub("<.*?>", "", name).strip()


def get_teamset_workbook(username: str, password: str, course_id: int) -> Workbook:
    session = requests.session()
    # get csrf token
    login_response = session.get("https://handins.ccs.neu.edu/login")
    soup = BeautifulSoup(login_response.content, "html.parser")
    csrf_param = soup.find("meta", {"name": "csrf-param"})["content"]
    csrf_token = soup.find("meta", {"name": "csrf-token"})["content"]
    # login using token
    session.post(
        "https://handins.ccs.neu.edu/login",
        {
            "user[username]": username,
            "user[password]": password,
            csrf_param: csrf_token,
        },
    )
    # get workbook
    teamset_response = session.get(
        f"https://handins.ccs.neu.edu/courses/{course_id}/teamsets/export.xlsx"
    )
    return load_workbook(BytesIO(teamset_response.content), read_only=True)


def parse(workbook: Workbook, worksheet_name: str) -> List[Team]:
    return [
        Team(
            team_id=row[0].value,
            active=row[1].value == "Yes",
            lab_section=int(row[4].value) if row[4].value else None,
            member1=without_email(row[6].value),
            member2=without_email(row[7].value) if row[7].value else None,
            member3=without_email(row[8].value) if row[8].value else None,
        )
        # skip title row
        for row in workbook[worksheet_name].iter_rows(min_row=2, max_col=9)
    ]


def main(
    course_id: int,
    lab_section: int,
    offset: int,
    sheet_name: Optional[str],
    latest: bool,
    staff: List[str],
    username: str,
    password: str,
):
    # fetch workbook
    workbook = get_teamset_workbook(username, password, course_id)
    if latest:
        sheet_name = workbook.sheetnames[-1]

    while not sheet_name:
        print("Select sheet number:")
        for index, sheetname in enumerate(workbook.sheetnames):
            print(f"{index}. {sheetname}")
        try:
            sheet_name = workbook.sheetnames[int(input())]
        except:
            print("Invalid selection.")

    # parse teams from workbook
    teams: List[Team] = sorted(parse(workbook, sheet_name))
    my_teams: List[Team] = [
        team for team in teams if team.lab_section == lab_section and team.active
    ]
    staff_assignments: List[Tuple[int, str]] = list(
        zip(range(len(staff)), staff[offset:] + staff[:offset])
    )

    # split teams into (almost) even groups
    grouped_teams: List[List[Team]] = [list(l) for l in divide(len(staff), my_teams)]
    # assign teams to a group number and find their min and max team ids
    team_assignments: List[Tuple[Tuple[int, str], int, int, List[Team]]] = list(
        zip(
            staff_assignments,
            map(lambda teams: min(teams).team_id, grouped_teams),
            map(lambda teams: max(teams).team_id, grouped_teams),
            grouped_teams,
        )
    )

    # write student mappings
    with open("students.csv", "w") as file:
        csvwriter = csv.writer(file)
        csvwriter.writerow(
            [
                "Group",
                "Team",
                "Staff",
                "Member 1",
                "Member 2",
                "Member 3",
                "Teams Channel Name",
            ]
        )
        for (group, staff_name), _, __, teams in team_assignments:
            for team in teams:
                csvwriter.writerow(
                    [
                        group,
                        team.team_id,
                        staff_name,
                        team.member1,
                        team.member2,
                        team.member3,
                        team.ms_team_channel,
                    ]
                )

    # write staff mappings
    with open("staff.csv", "w") as file:
        csvwriter = csv.writer(file)
        csvwriter.writerow(
            ["Group", "Start", "End", "Staff", "Teams Channel Name", "Count"]
        )
        for (group, staff_name), min_id, max_id, teams in team_assignments:
            csvwriter.writerow(
                [
                    group,
                    min_id,
                    max_id,
                    staff_name,
                    f"Groups ({min_id}-{max_id}) - {staff_name}",
                    len(teams),
                ]
            )


if __name__ == "__main__":
    # parse cli args
    parser = argparse.ArgumentParser(
        description="Bottlenose staff lab pairing utility",
        epilog=(
            "The KHOURY_USERNAME and KHOURY_PASSWORD environment "
            "variables can be set to fully automate this script"
        ),
    )
    parser.add_argument("course", type=int, help="ID of the course")
    parser.add_argument("lab", type=int, help="CRN of lab section")
    parser.add_argument("staff", nargs="+", help="Names of staff members")
    parser.add_argument(
        "-o", "--offset", type=int, default=0, help="Used to rotate staff members"
    )
    parser.add_argument(
        "-s", "--sheet", type=str, help="Name of sheet to read team data"
    )
    parser.add_argument(
        "--latest", action="store_true", help="Uses the latest teamsets"
    )
    args = parser.parse_args()

    main(
        course_id=args.course,
        lab_section=args.lab,
        offset=args.offset,
        sheet_name=args.sheet,
        latest=args.latest,
        staff=args.staff,
        username=os.getenv("KHOURY_USERNAME", input("Khoury username: ")),
        password=os.getenv("KHOURY_PASSWORD", getpass("Khoury password: ")),
    )
